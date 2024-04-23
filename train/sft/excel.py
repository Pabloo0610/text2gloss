import openpyxl

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()

# 获取活动的工作表
worksheet = workbook.active

# 打开TXT文件并读取内容
with open('yfs_10000_gt.txt', 'r', encoding='utf-8') as file:
    lines = file.readlines()

# 将TXT文件的每一行前半句写入Excel工作表的第一列,后半句写入第二列
for row_num, line in enumerate(lines, start=1):
    line_parts = line.split('Assistant: ')  # 以逗号分隔前后半句
    line_parts[1] = line_parts[1].split('Gt:')[0]
    #line_parts[1].strip('</s>\n')
    print(line_parts[1])
    if len(line_parts) == 2:
        worksheet.cell(row=row_num, column=1, value=line_parts[0].strip())
        worksheet.cell(row=row_num, column=2, value=line_parts[1].strip().rstrip('</s>\n'))
    else:
        worksheet.cell(row=row_num, column=1, value=line.strip())

# 保存Excel工作簿
workbook.save('output_10000.xlsx')
print("TXT文件已成功写入Excel工作簿!")